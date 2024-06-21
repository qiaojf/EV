import numpy as np
import pandas as pd
import math,copy,time
from openpyxl import load_workbook


class EV:#定义EV
    def __init__(self,id,*args):#args[0]是通勤时长，args[1]是通勤时间
        # print(args)
        args = args[0]
        # print(args)
        self.id = id #编号
        self.power = 51.4  #额定功率
        self.eit = 0.95 #充放电效率η 
        self.soc_min = 0    #最小电量比
        self.soc_max = 0.9      #最大电量比
        self.power_min = -50    #最大放电功率
        self.power_max = 50     #最大充电功率
        
        self.day1_go = args[1][1] if id <= 40 else args[1][0]    #第一天离开商业区时间
        self.day1_go_time = args[0][1] if id <= 40 else args[0][0]  #第一天回家路上所用时间
        self.day1_come = args[1][0] if id <= 40 else args[1][1]      #第一天到达商业区时间
        self.day1_come_time = args[0][0] if id <= 40 else args[0][1]     #第一天来商业区路上所用时间
        self.day1_reach_soc = self.power*self.soc_max-9.6*self.day1_come_time/60     #第一天到达商业区的soc
        self.day2_go = args[1][3] if id <= 40 else args[1][2]    #第二天离开商业区时间
        self.day2_go_time = args[0][3] if id <= 40 else args[0][2]   #第二天回家路上所用时间
        self.day2_come = args[1][2] if id <= 40 else args[1][3]    #第二天到达商业区时间
        self.day2_come_time = args[0][2] if id <= 40 else args[0][3]     #第二天来商业区路上所用时间
        self.day2_reach_soc = self.power*self.soc_max-9.6*self.day2_come_time/60    #第二天到达商业的soc
        self.min_cap = self.get_min_cap(max(args[0]))      #保证回家最小电量

        self.curr_charge_time = 0      #当前充放电时间
        self.is_here = 1 if id > 40 else 0        #当前是否在商业区1表示在
        self.ecl = 0   #充放电次数
        self.last_action = 0    #上一次是充电还是放电1表示充电-1表示放电
        self.last_cap = 0   #上次充放电量
        self.curr_soc = 0.5  #当前电量百分比
        self.curr_dod = 0.4  #当前放电深度
        self.curr_cap = self.power*self.curr_soc     #当前电量
        self.curr_power = 0     #当前充放电功率
        self.cap = self.power
        self.cost = 0
        self.reward = 0
        self.t_reward = 0
        
    def cal_5min_cap(self,power):#5min的充放电量
        cap = power*self.eit*5/60
        return cap

    def get_min_cap(self,time_len):#路上耗电量
        min_cap = 9.6*time_len/60
        return min_cap

    def cal_dod(self, p):#计算DOD
        # dod = p/self.curr_cap
        dod = (self.power*self.soc_max-self.curr_cap)/self.cap
        return dod

    def cal_ecl(self,dod):#计算ecl
        dod = abs(dod)
        a = 2807
        b = 0.02842
        c = -7.658
        d = 328.1
        f = -1597
        ecl = a*math.exp(b*dod)+c*dod**2+d*dod+f
        return ecl

    def cal_cap(self,soc,c_rate,dod,ecl):#计算剩余容量
        a = (6.2*soc/90+0.093)*2
        b = (0.98+0.01741*soc/20)*(-0.6045/(soc**24)-5.512*(10**-4))*(soc/20)
        c = (22-a)*2
        if 0 < soc <= 0.1:
            d = -6.620*10**-6
        elif 0.1 < soc <= 0.2:
            d = -3.210*10**-6
        elif 0.2 < soc <= 0.3:
            d = -2.410*10**-6
        elif 0.3 < soc <= 0.4:
            d = -3.700*10**-6
        elif 0.4 < soc <= 0.5:
            d = -5.000*10**-6
        elif 0.5 < soc <= 0.6:
            d = -2.550*10**-6
        elif 0.6 < soc <= 0.7:
            d = -0.100*10**-6
        elif 0.7 < soc <= 0.9:
            d = -0.010*10**-6
        else:
            d = -0.010*10**-6
        cap = a*math.exp(b*ecl)+c*math.exp(d*ecl)
        return cap

    def get_cost(self, cap,ecl,p):
        p = abs(p)
        w = 1
        c_rate = 2
        c_ev = 30840
        t = 1/12
        cost = -c_ev*p*t/(w*cap*self.eit**2*ecl)
        return cost

class Env:
    def __init__(self,num):#初始化环境
        self.ev_list = self.generate_ev(num)
        self.num_list_hour = self.get_num_hours()
        self.num_list_5mins = self.get_num_min()
        self.elec_price_list = self.get_elec_price()
        self.total_power_list = np.around(self.get_total_power(),decimals=2)
        self.curr_price = 0
        self.next_price = self.elec_price_list[0]
        self.curr_num = 0
        self.next_num = self.num_list_5mins[0]
        self.curr_power = 0
        self.next_power = self.total_power_list[0]
        self.resource = self.get_resources()

        self.state = self.init_state()
        self.actions_dim = num
        self.wb = load_workbook('程序数据集/my_data.xlsx')

    def init_state(self):#初始化环境的状态
        state = np.zeros((60,30))
        for ev in self.ev_list:
            id = self.ev_list.index(ev)
            state[id][0] = ev.id
            state[id][1] = ev.power
            state[id][2] = ev.eit
            state[id][3] = ev.soc_min
            state[id][4] = ev.soc_max
            state[id][5] = ev.power_min
            state[id][6] = ev.power_max
            state[id][7] = ev.day1_go
            state[id][8] = ev.day1_go_time
            state[id][9] = ev.day1_come
            state[id][10] = ev.day1_come_time
            state[id][11] = ev.day1_reach_soc
            state[id][12] = ev.day2_go
            state[id][13] = ev.day2_go_time
            state[id][14] = ev.day2_come
            state[id][15] = ev.day2_come_time
            state[id][16] = ev.day2_reach_soc
            state[id][17] = ev.min_cap
            state[id][18] = ev.curr_charge_time
            state[id][19] = ev.is_here
            state[id][20] = ev.ecl
            state[id][21] = ev.last_action
            state[id][22] = ev.curr_soc
            state[id][23] = ev.curr_dod
            state[id][24] = ev.curr_cap
            state[id][25] = ev.curr_power
            state[id][26] = self.curr_price
            state[id][27] = self.curr_power
            state[id][28] = self.curr_num
            state[id][29] = ev.last_cap
            state[id] = np.around(state[id],decimals=2)

        return state

    def write_excel(self,state,total_step):

        # 遍历所有工作表
        for sheet in self.wb.worksheets:
            if sheet.title == 'soc':
                for col in sheet.iter_cols():
                    sheet[col[0].column_letter+str(total_step+1)] = state[col[0].value-1,22]
            elif sheet.title == 'dod':
                for col in sheet.iter_cols():
                    sheet[col[0].column_letter+str(total_step+1)] = state[col[0].value-1,23]
            elif sheet.title == 'cap':
                for col in sheet.iter_cols():
                    sheet[col[0].column_letter+str(total_step+1)] = self.ev_list[col[0].value-1].cap
            elif sheet.title == 'ecl':
                for col in sheet.iter_cols():
                    sheet[col[0].column_letter+str(total_step+1)] = state[col[0].value-1,20]
            elif sheet.title == 'curr_power':
                for col in sheet.iter_cols():
                    sheet[col[0].column_letter+str(total_step+1)] = state[col[0].value-1,25]
            elif sheet.title == 'is_here':
                for col in sheet.iter_cols():
                    sheet[col[0].column_letter+str(total_step+1)] = state[col[0].value-1,19]
            elif sheet.title == 'curr_cap':
                for col in sheet.iter_cols():
                    sheet[col[0].column_letter+str(total_step+1)] = state[col[0].value-1,24]
            elif sheet.title == 'cost':
                for col in sheet.iter_cols():
                    sheet[col[0].column_letter+str(total_step+1)] = self.ev_list[col[0].value-1].cost
            elif sheet.title == 'reward':
                for col in sheet.iter_cols():
                    sheet[col[0].column_letter+str(total_step+1)] = self.ev_list[col[0].value-1].reward
            elif sheet.title == 't_reward':
                for col in sheet.iter_cols():
                    sheet[col[0].column_letter+str(total_step+1)] = self.ev_list[col[0].value-1].t_reward
        #保存工作表
        # self.wb.save('程序数据集/my_data.xlsx')


    def step(self,action,state,total_step):#更新每辆车的数据

        self.write_excel(state,total_step)#这行代码取消注释就会往表格写入soc等数据

        next_state = copy.deepcopy(state)
        self.curr_price = self.next_price
        self.next_price = self.elec_price_list[total_step] if total_step < 576 else 0
        self.curr_power = -self.next_power
        self.next_power = self.total_power_list[total_step] if total_step < 576 else 0
        self.curr_num = self.next_num
        self.next_num = self.num_list_5mins[total_step] if total_step < 576 else 0
        unit_power = self.curr_power/np.sum(action) if np.sum(action) != 0 else 0
        print('power:',unit_power)
        resource = self.resource[total_step-1]
        count = len([ev for ev in self.ev_list if ev.is_here])
        resouce_value = self.curr_price*resource/count

        cost_list = []
        for ev in self.ev_list:
            id = self.ev_list.index(ev)
            if ev.is_here:#EV在电网
                ev.last_cap = ev.cal_5min_cap(action[id]*unit_power)   #当前动作充电量
                ev.curr_power = action[id]*unit_power
                ev.ecl = ev.cal_ecl(ev.curr_dod)
                ev.curr_cap = ev.curr_cap + ev.last_cap
                ev.curr_cap = ev.curr_cap if ev.curr_cap < ev.power*ev.soc_max else ev.power*ev.soc_max
                ev.curr_cap = ev.curr_cap if ev.curr_cap > ev.min_cap else ev.min_cap
                # ev.curr_cap = evc if ev.min_cap < evc < ev.power*ev.soc_max else ev.curr_cap - ev.last_cap
                ev.curr_dod = ev.cal_dod(ev.last_cap)
                ev.curr_soc = ev.curr_cap/ev.power 
                ev.last_action = action[id]
                ev.cap = ev.cal_cap(ev.curr_soc,2,ev.curr_dod,ev.ecl)
                ev.cost = ev.get_cost(ev.cap,ev.ecl,ev.last_cap)
                ev.reward = self.curr_price*ev.last_cap
                ev.t_reward = ev.cost + ev.reward + resouce_value
                
                # next_state[id][18] = ev.curr_charge_time + 5 
                next_state[id][19] = ev.is_here
                next_state[id][20] = ev.ecl
                next_state[id][21] = ev.last_action
                next_state[id][22] = ev.curr_soc
                next_state[id][23] = ev.curr_dod
                next_state[id][24] = ev.curr_cap
                next_state[id][25] = ev.curr_power
                next_state[id][26] = self.curr_price
                next_state[id][27] = self.curr_power
                next_state[id][28] = self.curr_num
                next_state[id][29] = ev.last_cap
                if ev.curr_soc > 0.2:#奖励soc
                    if (action[id] < 0 and unit_power > 0) or (action[id] > 0 and unit_power < 0):
                        cost_list.append(ev.cost)
                next_state[id] = np.around(next_state[id],decimals=2)
            else:#EV不在电网
                max_cap = ev.power*ev.soc_max
                if ev.id <= 40:
                    if total_step == 1 or total_step == 493 + ev.day2_go//5:
                        road_cost = ev.get_min_cap(ev.day1_come_time)
                        ev.curr_cap = max_cap-road_cost
                        ev.curr_dod = road_cost/ev.cap
                        ev.ecl = ev.cal_ecl(ev.curr_dod)
                        ev.curr_soc = ev.curr_cap/ev.power
                    elif total_step == 205 + ev.day1_go//5:
                        road_cost = ev.get_min_cap(ev.day2_come_time)
                        ev.curr_cap = max_cap-road_cost
                        ev.curr_dod = road_cost/ev.cap
                        ev.ecl = ev.cal_ecl(ev.curr_dod)
                        ev.curr_soc = ev.curr_cap/ev.power
                else:
                    if total_step == 85 + ev.day1_go//5:
                        road_cost = ev.get_min_cap(ev.day1_come_time)
                        ev.curr_cap = max_cap-road_cost
                        ev.curr_dod = road_cost/ev.cap
                        ev.ecl = ev.cal_ecl(ev.curr_dod)
                        ev.curr_soc = ev.curr_cap/ev.power
                    elif total_step == 372 + ev.day2_go//5:
                        road_cost = ev.get_min_cap(ev.day2_come_time)
                        ev.curr_cap = max_cap-road_cost
                        ev.curr_dod = road_cost/ev.cap
                        ev.ecl = ev.cal_ecl(ev.curr_dod)
                        ev.curr_soc = ev.curr_cap/ev.power
                ev.last_cap = 0   #当前动作充电量
                ev.curr_power = 0
                ev.last_action = action[id]
                ev.cap = ev.power
                ev.cost = 0
                ev.reward = 0
                ev.t_reward = 0

                next_state[id][19] = ev.is_here
                next_state[id][20] = ev.ecl
                next_state[id][21] = ev.last_action
                next_state[id][22] = ev.curr_soc
                next_state[id][23] = ev.curr_dod
                next_state[id][24] = ev.curr_cap
                next_state[id][25] = ev.curr_power
                next_state[id][26] = self.curr_price
                next_state[id][27] = self.curr_power
                next_state[id][28] = self.curr_num
                next_state[id][29] = ev.last_cap

        cost = sum(cost_list)
        reward = self.curr_price*self.curr_power + cost + self.curr_price*resource
        # if (self.curr_power > 0 and unit_power > 0) or (self.curr_power < 0 and unit_power < 0):
        #     reward = abs(self.curr_price*self.curr_power) + cost
        # else:#采取动作以后的奖励计算reward = 充放电收益+消纳新能源收益+电车损耗
        #     reward = -abs(self.curr_price*self.curr_power) + cost
        print('reward:',reward)
        done = 0
        return next_state,reward,done

    def get_resources(self):
        resource_list = []
        resource_list = pd.read_excel('./程序数据集/商业微电网新能源出力5min.xlsx',sheet_name='Sheet1')
        return resource_list['新能源']

    def get_total_power(self):
        total_power_list = []
        total_power_list = pd.read_excel('./程序数据集/各单元出力.xlsx',sheet_name='snd_power')
        return total_power_list['电车功率']

    def get_num_hours(self):
        ev_num = pd.read_excel('./程序数据集/商业微电网通勤数据.xlsx',sheet_name='全时段微网车辆数记录')
        return ev_num.values

    def get_num_min(self):
        ev_num = []
        for num in self.num_list_hour:
            for i in range(12):
                ev_num.append(num)
        return ev_num

    def action_sample(self):#随机生成动作
        action = np.random.uniform(-1.0,1.0,60)
        action = np.around(action,decimals=2)
        # print(action)
        return action
    
    def update_ev(self,step):
        for ev in self.ev_list:
            if ev.id <= 40:#更新白班车辆是否在现场，以及到达时电量
                if 84 + ev.day1_come//5+1 < step <= 204 + ev.day1_go//5:
                    ev.is_here = 1
                elif 372 + ev.day2_come//5+1 < step <= 492 + ev.day2_go//5:
                    ev.is_here = 1
                else:
                    ev.is_here = 0
                if 84 + ev.day1_come//5+2 == step:
                    ev.curr_cap = ev.power*ev.soc_max - 9.6*ev.day1_come_time/60
                elif 372 + ev.day2_come//5+2 == step:
                    ev.curr_cap = ev.power*ev.soc_max - 9.6*ev.day2_come_time/60
            else:       #更新夜班车辆是否在现场，以及到达时电量
                if 84 + ev.day1_go//5 > step:
                    ev.is_here = 1
                elif 204 + ev.day1_come//5+1 < step <= 372 + ev.day2_go//5:
                    ev.is_here = 1
                elif step > 492 + ev.day2_come//5+1:
                    ev.is_here = 1
                else:
                    ev.is_here = 0
                if 204 + ev.day1_come//5+2 == step:
                    ev.curr_cap = ev.power*ev.soc_max - 9.6*ev.day1_come_time/60
                elif step == 492 + ev.day2_come//5+2:
                    ev.curr_cap = ev.power*ev.soc_max - 9.6*ev.day2_come_time/60
            ev.curr_cap = round(ev.curr_cap,2)

    def check_action(self,action):#对输出动作进行检查处理
        for ev in self.ev_list:
            if not ev.is_here:#不在的车辆充电功率设为0
                action[self.ev_list.index(ev)] = 0
            else:#如果充电后当前电量大于最大电量，或者小于最小电量
                cap = ev.cal_5min_cap(action[self.ev_list.index(ev)]*50)
                cap = round(cap,2)
                if ev.curr_cap + cap > ev.power*ev.soc_max:
                    action[self.ev_list.index(ev)] = -action[self.ev_list.index(ev)]
                elif ev.curr_cap + cap < ev.min_cap:
                    action[self.ev_list.index(ev)] = -action[self.ev_list.index(ev)]
                elif ev.curr_cap - cap > ev.power*ev.soc_max:
                    action[self.ev_list.index(ev)] = -action[self.ev_list.index(ev)]
                elif ev.curr_cap - cap < ev.min_cap:
                    action[self.ev_list.index(ev)] = -action[self.ev_list.index(ev)]

        unit_power = self.next_power/np.sum(action) if np.sum(action) != 0 else 0
        if unit_power > 50 or unit_power < -50:
            return (False,action)
        # else:#如果充电后当前电量大于最大电量，或者小于最小电量
        #     un_p = 50 if unit_power >= 0 else -50
        #     for ev in self.ev_list:
        #         cap = ev.cal_5min_cap(action[self.ev_list.index(ev)]*un_p)
        #         if ev.curr_cap + cap > ev.power*ev.soc_max:
        #             action[self.ev_list.index(ev)] = -action[self.ev_list.index(ev)]
        #         elif ev.curr_cap + cap < ev.min_cap:
        #             action[self.ev_list.index(ev)] = -action[self.ev_list.index(ev)]
        # unit_power = self.next_power/np.sum(action) if np.sum(action) != 0 else 0
        # if unit_power > 50 or unit_power < -50:
        #     return (False,action)
        return (True,action)

    def get_elec_price(self):
        price = pd.read_excel('./程序数据集/实时电价.xlsx')
        price_list =[]
        for p in price['实时电价(美元/h)'].values:
            for i in range(12):
                price_list.append(p)
        return price_list

    def generate_ev(self,num=60):#生成ev列表
        ev_list = []
        ev_commute_len = pd.read_excel('./程序数据集/商业微电网通勤数据.xlsx',sheet_name='电车通勤时长')
        ev_day_time = pd.read_excel('./程序数据集/商业微电网通勤数据.xlsx',sheet_name='白班到达、离开时间')
        ev_night_time = pd.read_excel('./程序数据集/商业微电网通勤数据.xlsx',sheet_name='夜班到达、离开时间')
        for i in range(1,num+1):#生成EV
            args = []
            args.append(ev_commute_len.values[i-1][1:].tolist())
            if i <= 40:
                args.append(ev_day_time.values[i-1][1:].tolist())
            else:
                args.append(ev_night_time.values[i-41][:].tolist())
            # print(i,args)
            ev_list.append(EV(i,args))
        return ev_list






